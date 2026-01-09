"""
Views for generating and managing reports from Okta data.

This module contains views for report generation, scheduling, and management.
"""
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings

import logging
import datetime
import json
import csv
import io
import uuid

from core.services.database import DatabaseService
from apps.analytics.services.metrics_service import get_metrics_data

# Optional format libraries
try:
    import openpyxl
    from openpyxl.styles import Font
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

logger = logging.getLogger(__name__)



class ReportDashboardView(LoginRequiredMixin, TemplateView):
    """
    View for displaying and managing reports.
    Shows available reports, scheduled reports, and report history.
    """
    template_name = 'traffic_analysis/reports/report_dashboard.html'
    login_url = '/login/'
    
    def dispatch(self, request, *args, **kwargs):
        # Store request in the instance for later use in get_context_data
        self.request = request
        return super().dispatch(request, *args, **kwargs)
    
    @method_decorator(cache_page(60 * 5))  # Cache for 5 minutes
    def get(self, request, *args, **kwargs):
        """Handle GET requests: instantiate a template response"""
        return super().get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        """Add report data to context for dashboard display"""
        context = super().get_context_data(**kwargs)
        
        # Use the nonce from the request (set by SecurityHeadersMiddleware)
        # instead of generating a new one
        if hasattr(self.request, 'nonce'):
            context['nonce'] = self.request.nonce
        
        # Add sample data for initial template rendering
        # In a real implementation, these would be fetched from services
        context.update({
            'available_reports': [
                {'key': 'user_activity', 'name': 'User Activity Report', 'description': 'Summary of user login and app usage activity', 'type': 'user'},
                {'key': 'security_compliance', 'name': 'Security Compliance Report', 'description': 'Analysis of security policy compliance', 'type': 'security'},
                {'key': 'mfa_usage', 'name': 'MFA Usage Report', 'description': 'Detail of MFA enrollment and usage across organization', 'type': 'security'},
                {'key': 'application_usage', 'name': 'Application Usage Report', 'description': 'Summary of application access and usage', 'type': 'application'},
                {'key': 'admin_changes', 'name': 'Administrative Changes Report', 'description': 'Log of all administrative changes', 'type': 'admin'}
            ],
            'scheduled_reports': [],
            'recent_reports': []
        })
        
        return context

def flatten_dict(d, parent_key='', sep='_'):
    """Flatten nested dictionaries for CSV export."""
    items = []
    for k, v in d.items():
        new_key = f'{parent_key}{sep}{k}' if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, (list, tuple)):
            items.append((new_key, str(v)))
        else:
            items.append((new_key, v))
    return dict(items)

def to_csv(data_dict):
    """Convert report data to CSV."""
    output = io.StringIO()
    flat_data = flatten_dict(data_dict)
    writer = csv.DictWriter(output, fieldnames=flat_data.keys())
    writer.writeheader()
    writer.writerow(flat_data)
    return output.getvalue().encode('utf-8')

def to_xlsx(data_dict, title='Report'):
    """Convert report data to XLSX."""
    if not HAS_OPENPYXL:
        raise Exception('openpyxl not installed')
    
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Report'
    
    # Add title
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    
    # Add data
    row = 3
    flat_data = flatten_dict(data_dict)
    for key, value in flat_data.items():
        ws[f'A{row}'] = key
        ws[f'B{row}'] = str(value) if not isinstance(value, (int, float)) else value
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def to_pdf(data_dict, title='Report'):
    """Convert report data to PDF."""
    if not HAS_REPORTLAB:
        raise Exception('reportlab not installed')
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Add title
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1a1a1a'), spaceAfter=20)
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Add metadata
    flat_data = flatten_dict(data_dict)
    table_data = [['Field', 'Value']]
    for key, value in list(flat_data.items())[:20]:  # Limit to first 20 rows
        table_data.append([str(key), str(value)[:100]])
    
    table = Table(table_data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4B77BE')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    story.append(table)
    
    doc.build(story)
    output.seek(0)
    return output.getvalue()

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_report(request):
    try:
        payload = request.data or {}
        report_key = payload.get('key') or payload.get('report') or ''
        days = int(payload.get('days', 30))
        fmt = (payload.get('format') or 'JSON').upper()

        metrics = get_metrics_data(days)
        now_str = datetime.datetime.now(datetime.timezone.utc).isoformat()
        
        # Map report keys to friendly names
        report_names = {
            'user_activity': 'User Activity Report',
            'security_compliance': 'Security Compliance Report',
            'mfa_usage': 'MFA Usage Report',
            'application_usage': 'Application Usage Report',
            'admin_changes': 'Administrative Changes Report'
        }

        # Compose report data based on key
        data = {}
        if report_key == 'user_activity':
            data = {
                'auth_activity': metrics.get('auth_activity'),
                'usage_by_app': metrics.get('usage_by_app'),
                'usage_by_device': metrics.get('usage_by_device'),
                'peak_usage_hour': metrics.get('peak_usage_hour'),
            }
        elif report_key == 'security_compliance':
            data = {
                'auth_success_rate': metrics.get('auth_success_rate'),
                'failed_logins_count': metrics.get('failed_logins_count'),
                'failed_logins_change': metrics.get('failed_logins_change'),
                'mfa_usage_rate': metrics.get('mfa_usage_rate'),
            }
        elif report_key == 'mfa_usage':
            data = {
                'auth_methods': metrics.get('auth_methods'),
                'mfa_usage_rate': metrics.get('mfa_usage_rate'),
                'mfa_rate_change': metrics.get('mfa_rate_change'),
            }
        elif report_key == 'application_usage':
            data = {
                'usage_by_app': metrics.get('usage_by_app'),
                'usage_by_location': metrics.get('usage_by_location'),
            }
        elif report_key == 'admin_changes':
            # Approximate admin changes via actor type and event patterns
            try:
                db = DatabaseService()
                coll = db.get_collection(settings.MONGODB_SETTINGS.get('db', 'OktaDashboardDB'), 'okta_logs')
                start = datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(days=days)
                start_str = start.isoformat().replace('+00:00', 'Z')
                now_iso = datetime.datetime.now(datetime.timezone.utc).isoformat().replace('+00:00', 'Z')
                admin_count = coll.count_documents({
                    'actor.type': {'$in': ['OktaAdmin', 'SystemAdmin']},
                    'published': {'$gte': start_str, '$lt': now_iso}
                })
                data = {'admin_events_count': admin_count}
            except Exception:
                data = {'admin_events_count': 0}
        else:
            return Response({'error': 'Unknown report key'}, status=status.HTTP_400_BAD_REQUEST)

        # Store report history in session
        history = request.session.get('report_history', [])
        history.insert(0, {
            'report_key': report_key,
            'report_name': report_names.get(report_key, report_key),
            'format': fmt,
            'days': days,
            'generated_at': now_str,
            'user': request.user.username if hasattr(request.user, 'username') else 'Unknown'
        })
        # Keep only last 50 reports
        request.session['report_history'] = history[:50]
        request.session.modified = True

        # Convert to requested format
        report_title = f'{report_key.replace("_", " ").title()} Report'
        if fmt == 'JSON':
            response_data = {
                'key': report_key,
                'generated_at': now_str,
                'format': fmt,
                'days': days,
                'data': data
            }
            return Response(response_data, status=status.HTTP_200_OK)
        elif fmt == 'CSV':
            csv_data = to_csv(data)
            response = HttpResponse(csv_data, content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{report_key}_{days}d.csv"'
            return response
        elif fmt == 'XLSX':
            if not HAS_OPENPYXL:
                return Response({'error': 'XLSX export not available'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)
            xlsx_data = to_xlsx(data, report_title)
            response = HttpResponse(xlsx_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{report_key}_{days}d.xlsx"'
            return response
        else:
            return Response({'error': f'Unsupported format: {fmt}'}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        logger.error(f'Error generating report: {e}', exc_info=True)
        return Response({'error': 'Failed to generate report', 'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def configure_report(request):
    try:
        payload = request.data or {}
        report_key = payload.get('key') or ''
        config = {
            'format': (payload.get('format') or 'PDF').upper(),
            'days': int(payload.get('days', 30)),
            'recipients': payload.get('recipients', ''),
        }
        if not report_key:
            return Response({'error': 'Missing report key'}, status=status.HTTP_400_BAD_REQUEST)
        # Persist to session for now (no DB schema yet)
        session_configs = request.session.get('report_configs', {})
        session_configs[report_key] = config
        request.session['report_configs'] = session_configs
        request.session.modified = True
        return Response({'key': report_key, 'config': config, 'status': 'saved'}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f'Error configuring report: {e}', exc_info=True)
        return Response({'error': 'Failed to configure report', 'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_report_history(request):
    """Get report generation history from session."""
    try:
        history = request.session.get('report_history', [])
        return Response({'reports': history}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f'Error fetching report history: {e}', exc_info=True)
        return Response({'error': 'Failed to fetch report history', 'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def manage_templates(request):
    """Create or list report templates."""
    try:
        if request.method == 'GET':
            # Get all templates from session
            templates = request.session.get('report_templates', [])
            return Response({'templates': templates}, status=status.HTTP_200_OK)
        
        elif request.method == 'POST':
            # Create new template
            payload = request.data or {}
            name = payload.get('name', '').strip()
            description = payload.get('description', '').strip()
            report_type = payload.get('reportType', '')
            format_type = payload.get('format', 'JSON').upper()
            days = int(payload.get('days', 30))
            
            if not name or not report_type:
                return Response({'error': 'Name and report type are required'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Generate unique ID
            template_id = str(uuid.uuid4())
            
            template = {
                'id': template_id,
                'name': name,
                'description': description,
                'report_type': report_type,
                'format': format_type,
                'days': days,
                'created_at': datetime.datetime.now(datetime.timezone.utc).isoformat(),
                'created_by': request.user.username if hasattr(request.user, 'username') else 'Unknown'
            }
            
            templates = request.session.get('report_templates', [])
            templates.append(template)
            request.session['report_templates'] = templates
            request.session.modified = True
            
            return Response({'template': template, 'status': 'created'}, status=status.HTTP_201_CREATED)
    except Exception as e:
        logger.error(f'Error managing templates: {e}', exc_info=True)
        return Response({'error': 'Failed to manage templates', 'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET', 'DELETE'])
@permission_classes([IsAuthenticated])
def template_detail(request, template_id):
    """Get or delete a specific template."""
    try:
        templates = request.session.get('report_templates', [])
        template = next((t for t in templates if t['id'] == template_id), None)
        
        if not template:
            return Response({'error': 'Template not found'}, status=status.HTTP_404_NOT_FOUND)
        
        if request.method == 'GET':
            return Response(template, status=status.HTTP_200_OK)
        
        elif request.method == 'DELETE':
            templates = [t for t in templates if t['id'] != template_id]
            request.session['report_templates'] = templates
            request.session.modified = True
            return Response({'status': 'deleted'}, status=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f'Error with template detail: {e}', exc_info=True)
        return Response({'error': 'Failed to process template', 'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
